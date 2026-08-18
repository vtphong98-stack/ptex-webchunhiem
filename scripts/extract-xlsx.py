"""Parse 12c1cn.xlsx into JSON for Mongo import. Does not write to the repo copy of the workbook."""

from __future__ import annotations

import json
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

SHEETS = {
    "TT1": {"role": "toTruong", "teamNumber": 1, "kind": "tt"},
    "TT2": {"role": "toTruong", "teamNumber": 2, "kind": "tt"},
    "TT3": {"role": "toTruong", "teamNumber": 3, "kind": "tt"},
    "TT4": {"role": "toTruong", "teamNumber": 4, "kind": "tt"},
    "LT": {"role": "lopTruong", "teamNumber": None, "kind": "lt"},
    "LPHT": {"role": "lopPhoHocTap", "teamNumber": None, "kind": "lpht"},
    "LPTT": {"role": "lopPhoTratTu", "teamNumber": None, "kind": "lptt"},
    "LPLD": {"role": "lopPhoLaoDong", "teamNumber": None, "kind": "lpld"},
    "LPPT": {"role": "lopPhoPhongTrao", "teamNumber": None, "kind": "lppt"},
    "ThuQuy": {"role": "thuQuy", "teamNumber": None, "kind": "thuquy"},
    "Tổng kết GVCN": {"role": "gvcn", "teamNumber": None, "kind": "gvcn"},
}


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    if text.startswith("="):
        return ""
    return text


def has_content(fields: dict[str, str]) -> bool:
    skip = {"week_range", "team_score", "total_income", "total_expense", "remaining"}
    return any(value.strip() for key, value in fields.items() if key not in skip)


def map_fields(kind: str, ws, row: int) -> dict[str, str]:
    g = lambda col: cell_text(ws.cell(row, col).value)
    if kind == "tt":
        return {
            "week_range": g(2),
            "not_prepared_names": g(3),
            "not_prepared_count": g(4),
            "no_homework_names": g(5),
            "no_homework_count": g(6),
            "disorder_names": g(7),
            "disorder_count": g(8),
            "late_names": g(9),
            "late_count": g(10),
            "violation_names": g(11),
            "violation_count": g(12),
            "absent_names": g(13),
            "absent_count": g(14),
            "good_points_names": g(15),
            "good_points_count": g(16),
            "participation_names": g(17),
            "participation_count": g(18),
        }
    if kind == "lt":
        return {
            "week_range": g(2),
            "notice_guild": g(3),
            "absent_student": g(4),
            "late_student": g(5),
            "violation_guild": g(6),
            "future_plan": g(7),
        }
    if kind == "lpht":
        return {
            "week_range": g(2),
            "good_points": g(3),
            "speaking": g(4),
            "teacher_reminded": g(5),
            "future_plan": g(6),
            "suggestions": g(7),
        }
    if kind == "lptt":
        return {
            "week_range": g(2),
            "disorder_not_sdb": g(3),
            "disorder_not_sdb_count": g(4),
            "disorder_sdb": g(5),
            "disorder_sdb_count": g(6),
            "social_media": g(7),
        }
    if kind == "lpld":
        return {
            "week_range": g(2),
            "cleaning_team": g(3),
            "chair_team": g(4),
            "monday": g(5),
            "tuesday": g(6),
            "wednesday": g(7),
            "thursday": g(8),
            "friday": g(9),
            "saturday": g(10),
            "late_duty": g(11),
            "feedback": g(12),
        }
    if kind == "lppt":
        return {
            "week_range": g(2),
            "campaign_name": g(3),
            "implementation_time": g(4),
            "progress": g(5),
            "assigned_students": g(6),
            "competition_date": g(7),
            "estimated_cost": g(8),
        }
    if kind == "thuquy":
        fields = {
            "week_range": g(2),
            "fee_per_student": g(3),
            "quantity_paid": g(4),
            "missing_students": g(5),
            "quantity_missing": g(6),
        }
        col = 7
        for index in range(1, 7):
            fields[f"expense_name_{index}"] = g(col)
            fields[f"expense_amount_{index}"] = g(col + 1)
            col += 2
        return fields
    return {"week_range": g(2), "summary": g(2) if kind != "gvcn" else g(2)}


def extract(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    reports = []
    for sheet_name, meta in SHEETS.items():
        ws = wb[sheet_name]
        for row in range(2, 37):
            week_raw = ws.cell(row, 1).value
            try:
                week_number = int(float(week_raw))
            except (TypeError, ValueError):
                continue
            if meta["kind"] == "gvcn":
                fields = {"summary": cell_text(ws.cell(row, 2).value)}
            else:
                fields = map_fields(meta["kind"], ws, row)
            if not has_content(fields):
                continue
            reports.append(
                {
                    "sheet": sheet_name,
                    "weekNumber": week_number,
                    "reporterRole": meta["role"],
                    "teamNumber": meta["teamNumber"],
                    "fields": fields,
                }
            )
    return {"source": str(path), "count": len(reports), "reports": reports}


def main() -> None:
    source = Path(sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\PhongPC\Downloads\12c1cn.xlsx")
    dest = Path(sys.argv[2] if len(sys.argv) > 2 else "scratch/xlsx_payload.json")
    dest.parent.mkdir(parents=True, exist_ok=True)
    payload = extract(source)
    dest.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({"source": payload["source"], "count": payload["count"], "out": str(dest)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
